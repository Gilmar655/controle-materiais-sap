from __future__ import annotations

import csv
import json
import math
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(sys.argv[1]).resolve()
DATA_DIR = ROOT / "data"
ASSETS_DIR = ROOT / "assets"

CANONICAL_FIELDS = [
    "projeto",
    "descricaoProjeto",
    "contrato",
    "codigoMaterial",
    "descricaoMaterial",
    "unidadeMedida",
    "quantidadeOrcada",
    "quantidadeBaixada",
    "reserva",
    "codigoMovimento",
    "tipoMovimento",
    "dataMovimento",
    "centro",
    "deposito",
    "documentoSap",
    "itemSap",
    "categoria",
    "materialOrigem",
    "classificacaoMcpse",
    "criterioOrdenacao",
    "pontoDescarga",
    "recebedor",
    "relevancia",
    "percentualFonteTotal",
    "percentualFonteItem",
]

HEADER_MAP = {
    "Definição do projeto": "projeto",
    "Material": "codigoMaterial",
    "Qtd Planej": "quantidadeOrcada",
    "Descrição": "descricaoMaterial",
    "Oper": "tipoMovimento",
    "Categ": "categoria",
    "Mtl Origem": "materialOrigem",
    "Qtd Real": "quantidadeBaixada",
    "Contrato": "contrato",
    "Classif. MCPSE": "classificacaoMcpse",
    "Crit Orden": "criterioOrdenacao",
    "Mov": "codigoMovimento",
    "Ponto Desc.": "pontoDescarga",
    "Item": "itemSap",
    "Reserva": "reserva",
    "Centro": "centro",
    "Dep": "deposito",
    "Recebedor": "recebedor",
    "Diagrama": "documentoSap",
    "Data da Necessidade": "dataMovimento",
    "Relev": "relevancia",
    "% Total": "percentualFonteTotal",
    "% do Item": "percentualFonteItem",
}

FIELD_DICTIONARY = [
    ("projeto", "Projeto", "Identificador do projeto ou Elemento PEP", "Definição do projeto; Projeto; Elemento PEP; PEP"),
    ("descricaoProjeto", "Descrição do Projeto", "Nome ou descrição do projeto", "Descrição do Projeto; Nome Projeto"),
    ("contrato", "Contrato", "Número ou referência do contrato associado", "Contrato; Número Contrato; Contrato SAP"),
    ("codigoMaterial", "Código do Material", "Código SAP do material", "Material; Código Material; Cod Material"),
    ("descricaoMaterial", "Descrição do Material", "Texto breve ou descrição do material", "Descrição; Texto breve material"),
    ("unidadeMedida", "Unidade de Medida", "Unidade utilizada na quantidade", "Unidade; UM; Unidade de Medida"),
    ("quantidadeOrcada", "Quantidade Orçada", "Quantidade planejada/orçada", "Qtd Planej; Qtd Planejada; Quantidade Orçada; Planejado"),
    ("quantidadeBaixada", "Quantidade Baixada", "Quantidade realizada/baixada", "Qtd Real; Qtd Realizada; Quantidade Baixada; Realizado"),
    ("reserva", "Reserva", "Número da reserva SAP", "Reserva; Número Reserva"),
    ("codigoMovimento", "Código do Movimento", "Código do movimento SAP", "Mov; Movimento; Código Movimento"),
    ("tipoMovimento", "Tipo de Movimento", "Operação ou tipo de movimento", "Oper; Tipo Movimento"),
    ("dataMovimento", "Data do Movimento", "Data da necessidade, reserva ou lançamento", "Data da Necessidade; Data Movimento; Data Lançamento"),
    ("centro", "Centro", "Centro SAP", "Centro"),
    ("deposito", "Depósito", "Depósito SAP", "Dep; Depósito"),
    ("documentoSap", "Documento SAP", "Documento ou diagrama SAP", "Diagrama; Documento SAP"),
    ("itemSap", "Item SAP", "Item do documento/reserva", "Item; Item SAP"),
]


def clean_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    return value


def as_number(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def load_records():
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    if "Materiais" not in workbook.sheetnames:
        raise RuntimeError("A aba 'Materiais' não foi encontrada na base de origem.")
    sheet = workbook["Materiais"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    mapped_headers = [HEADER_MAP.get(header, "") for header in headers]

    records = []
    for source_row, values in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in values):
            continue
        record = {field: "" for field in CANONICAL_FIELDS}
        for field, value in zip(mapped_headers, values):
            if field:
                record[field] = clean_value(value)
        record["quantidadeOrcada"] = as_number(record["quantidadeOrcada"])
        record["quantidadeBaixada"] = as_number(record["quantidadeBaixada"])
        record["_linhaOrigem"] = source_row
        records.append(record)
    return records, headers


def summarize(records):
    projects = {str(row["projeto"]).strip() for row in records if str(row["projeto"]).strip()}
    materials = {str(row["codigoMaterial"]).strip() for row in records if str(row["codigoMaterial"]).strip()}
    contracts = {str(row["contrato"]).strip() for row in records if str(row["contrato"]).strip()}
    movements = {str(row["codigoMovimento"]).strip() for row in records if str(row["codigoMovimento"]).strip()}
    reservations = {str(row["reserva"]).strip() for row in records if str(row["reserva"]).strip()}
    planned = sum(row["quantidadeOrcada"] for row in records)
    downloaded = sum(row["quantidadeBaixada"] for row in records)
    missing_contracts = sum(not str(row["contrato"]).strip() for row in records)
    negative_planned = sum(row["quantidadeOrcada"] < 0 for row in records)
    negative_downloaded = sum(row["quantidadeBaixada"] < 0 for row in records)
    duplicate_keys = Counter(
        (
            str(row["projeto"]).strip(),
            str(row["codigoMaterial"]).strip(),
            str(row["documentoSap"]).strip(),
            str(row["itemSap"]).strip(),
        )
        for row in records
    )
    duplicate_rows = sum(count for key, count in duplicate_keys.items() if all(key) and count > 1)
    return {
        "totalRegistros": len(records),
        "totalProjetos": len(projects),
        "totalMateriais": len(materials),
        "totalContratos": len(contracts),
        "totalMovimentos": len(movements),
        "totalReservas": len(reservations),
        "quantidadeOrcada": round(planned, 6),
        "quantidadeBaixada": round(downloaded, 6),
        "saldoNaoBaixado": round(planned - downloaded, 6),
        "percentualBaixa": round(downloaded / planned * 100, 6) if planned else 0,
        "registrosSemContrato": missing_contracts,
        "quantidadesOrcadasNegativas": negative_planned,
        "quantidadesBaixadasNegativas": negative_downloaded,
        "registrosDuplicados": duplicate_rows,
    }


def write_data(records, headers):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    summary = summarize(records)
    payload = {
        "metadata": {
            "titulo": "Controle de Materiais Orçados × Materiais Baixados — SAP",
            "origem": "SAP",
            "arquivo": SOURCE.name,
            "aba": "Materiais",
            "dataAtualizacao": "2026-07-24T12:00:00-03:00",
            "geradoEm": datetime.now().astimezone().isoformat(timespec="seconds"),
            "cabecalhosOriginais": headers,
            "resumoValidacao": summary,
        },
        "records": records,
    }
    json_text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    (DATA_DIR / "base-inicial.json").write_text(json_text, encoding="utf-8")
    (DATA_DIR / "base-embed.js").write_text(
        "window.BASE_INICIAL_EMBUTIDA=" + json_text + ";\n",
        encoding="utf-8",
    )
    with (DATA_DIR / "dicionario-campos.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(["Campo canônico", "Nome sugerido", "Descrição", "Aliases reconhecidos"])
        writer.writerows(FIELD_DICTIONARY)
    (DATA_DIR / "resumo-validacao.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return summary


def font(size, bold=False):
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()


def make_assets():
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    canvas = Image.new("RGBA", (520, 180), (255, 255, 255, 0))
    draw = ImageDraw.Draw(canvas)
    draw.rounded_rectangle((8, 8, 512, 172), radius=28, fill=(255, 255, 255, 242), outline=(210, 218, 232, 255), width=3)
    draw.rounded_rectangle((26, 30, 112, 116), radius=24, fill=(4, 119, 242, 255))
    draw.ellipse((68, 68, 128, 128), fill=(232, 24, 111, 235))
    draw.text((148, 38), "ENEL BRASIL", fill=(24, 32, 62, 255), font=font(32, True))
    draw.text((148, 88), "INSIRA A LOGOMARCA OFICIAL", fill=(91, 104, 129, 255), font=font(17))
    canvas.save(ASSETS_DIR / "logo-enel.png")

    icon = Image.new("RGBA", (128, 128), (15, 23, 42, 255))
    icon_draw = ImageDraw.Draw(icon)
    icon_draw.rounded_rectangle((14, 14, 114, 114), radius=28, fill=(4, 119, 242, 255))
    icon_draw.ellipse((55, 50, 110, 105), fill=(232, 24, 111, 235))
    icon_draw.text((27, 34), "M", fill=(255, 255, 255, 255), font=font(48, True))
    icon.save(ASSETS_DIR / "favicon.png")


if __name__ == "__main__":
    records, source_headers = load_records()
    result = write_data(records, source_headers)
    make_assets()
    print(json.dumps(result, ensure_ascii=False, indent=2))
